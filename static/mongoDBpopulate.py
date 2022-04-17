# creating index based on timestamp and symbol
from openpyxl import load_workbook, Workbook

def mongoDbpopulate(collection):
    print(collection.index_information())

    # Loading all the excel files
    wb = load_workbook("./data/BNBUSDT.xlsx",data_only=True)
    shBNB = wb["BNBUSDT"]

    wb = load_workbook("./data/BTCUSDT.xlsx",data_only=True)
    shBTC = wb["BTCUSDT"]

    wb = load_workbook("./data/ETHUSDT.xlsx", data_only=True)
    shETH = wb["ETHUSDT"]

    wb = load_workbook("./data/LTCUSDT.xlsx", data_only=True)
    shLTC = wb["LTCUSDT"]

    wb = load_workbook("./data/NEOUSDT.xlsx", data_only=True)
    shNEO = wb["NEOUSDT"]

    # Gather all the data of BNB table into map
    cryptoList = []
    sheetList = [
        {"sheet": shBNB, "symbol": "BNB"},
        {"sheet": shBTC, "symbol": "BTC"},
        {"sheet": shETH, "symbol": "ETH"},
        {"sheet": shLTC, "symbol": "LTC"},
        {"sheet": shNEO, "symbol": "NEO"}
    ]
    # unix	date	symbol	open	high	low	close	Volume BTC	Volume USDT	tradecount

    # Loop through every crypto sheet
    for dict in sheetList:
        symbol = dict["symbol"]
        for row in dict["sheet"].iter_rows(min_row=2):
            map = {} 
            map["timestamp"] = row[0].value
            map["date"] = row[1].value
            map["symbol"] = symbol
            map["open"] = row[3].value
            map["high"] = row[4].value
            map["low"] = row[5].value
            map["close"] = row[6].value
            map["VolumeBTC"] = row[7].value
            map["VolumeUSDT"] = row[8].value
            map["tradecount"] = row[9].value
            cryptoList.append(map)
        
        print(cryptoList[0])
        print(len(cryptoList))
        print()
        collection.insert_many(cryptoList)
        cryptoList = []
        
